"""산출물 — 콘솔 리포트 / HTML 검증 리포트 / 집행내역서 xlsx / 반납 기안 초안."""
from __future__ import annotations

import html
from typing import List, Optional

from .models import (ERROR, INFO, OK, WARNING, Draft, Project, Quote, Result,
                     TaxInvoice)
from .normalize import format_money as fm
from .normalize import korean_amount

__all__ = ["text_report", "html_report", "refund_draft", "settlement_rows",
           "write_xlsx"]

_MARK = {ERROR: "✕", WARNING: "!", INFO: "i", OK: "✓"}


# ------------------------------------------------------------------ 콘솔

def text_report(result: Result, project: Optional[Project] = None) -> str:
    s = result.summary()
    out: List[str] = []
    if project:
        out.append(f"■ {project.name}")
        if project.start and project.end:
            out.append(f"  집행기간 {project.start} ~ {project.end}")
    out.append(f"  교부 {fm(project.grant) if project else '—'}"
               f" / 기집행 {fm(result.prior_total)}"
               f" / 이번 건 {fm(result.this_amount)}"
               f" / 잔액 {fm(result.remaining)}")
    if result.line_name and result.line_allocated is not None:
        out.append(f"  세부항목 「{result.line_name}」 "
                   f"{fm(result.line_spent)} / {fm(result.line_allocated)}")
    out.append("")
    out.append(f"검증: 오류 {s['error']} · 경고 {s['warning']} · "
               f"참고 {s['info']} · 통과 {s['ok']}")
    out.append("-" * 66)
    for f in result.sorted_findings:
        out.append(f"{_MARK[f.severity]} {f.code}  {f.title}")
        if f.detail:
            out.append(f"        {f.detail}")
    out.append("-" * 66)
    out.append("확정 가능" if result.ok else "⚠ 오류 해결 전 확정 불가")
    return "\n".join(out)


# -------------------------------------------------------------- 집행내역서

def settlement_rows(project: Project, result: Result,
                    draft: Optional[Draft] = None,
                    quote: Optional[Quote] = None) -> List[list]:
    """집행내역서 본문 행. [연번, 집행일, 세부항목, 거래처, 사업자번호, 금액, 계약방법]"""
    rows: List[list] = []
    for i, p in enumerate(project.prior, 1):
        rows.append([i, str(p.spend_date or ""), p.line, p.vendor_name,
                     p.biz_no, p.amount, p.method])
    d = (draft.draft_date if draft else None) or (quote.quote_date if quote else None)
    vendor = ((quote.vendor.name if quote and quote.vendor else None)
              or (draft.vendor.name if draft and draft.vendor else None) or "")
    biz = ((quote.vendor.biz_no if quote and quote.vendor else None)
           or (draft.vendor.biz_no if draft and draft.vendor else None) or "")
    rows.append([len(rows) + 1, str(d or ""), result.line_name, vendor, biz,
                 result.this_amount, (draft.method if draft else "") or ""])
    return rows


def write_xlsx(path: str, project: Project, result: Result,
               draft: Optional[Draft] = None, quote: Optional[Quote] = None) -> str:
    """집행내역서를 xlsx로 저장한다. openpyxl 필요."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "집행내역서"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EFEFEF")

    ws.merge_cells("A1:G1")
    ws["A1"] = "목적사업비 집행내역서"
    ws["A1"].font = Font(size=15, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    period = f"{project.start} ~ {project.end}" if project.start and project.end else ""
    ws["A2"] = f"{project.name}    집행기간 {period}"
    ws["A2"].alignment = Alignment(horizontal="center")

    header = ["연번", "집행일", "세부항목", "거래처", "사업자등록번호", "집행액(원)", "계약방법"]
    ws.append([])
    ws.append(header)
    hrow = ws.max_row
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=hrow, column=c)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row in settlement_rows(project, result, draft, quote):
        ws.append(row)
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=ws.max_row, column=c)
            cell.border = border
            if c == 6:
                cell.number_format = "#,##0"

    ws.append([])
    for label, value in (("교부액", project.grant),
                         ("집행액 계", result.total_spent),
                         ("집행잔액 (반납 대상)", result.remaining)):
        ws.append(["", "", "", "", label, value, ""])
        ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
        c = ws.cell(row=ws.max_row, column=6)
        c.font = Font(bold=True)
        c.number_format = "#,##0"

    ws.append([])
    ws.append(["검증", f"오류 {len(result.errors)} · 경고 {len(result.warnings)}"
                       f" · 통과 {len(result.by_severity(OK))}"])
    for f in result.sorted_findings:
        if f.severity in (ERROR, WARNING):
            ws.append(["", f"{_MARK[f.severity]} {f.code} {f.title}", f.detail or ""])

    for col, width in zip("ABCDEFG", (6, 13, 22, 20, 18, 15, 22)):
        ws.column_dimensions[col].width = width

    wb.save(path)
    return path


# ------------------------------------------------------------ 반납 기안 초안

def refund_draft(project: Project, result: Result) -> Optional[str]:
    """집행잔액이 있으면 반납 기안 초안 텍스트를 만든다."""
    if result.remaining <= 0:
        return None
    rate = (result.total_spent / project.grant * 100) if project.grant else 0
    return (
        f"제목: {project.name} 목적사업비 집행잔액 반납\n\n"
        f"1. 관련: {project.grant_no or project.name} 목적사업비 교부\n"
        f"2. 위 사업의 집행을 완료하고 잔액을 아래와 같이 반납하고자 합니다.\n\n"
        f"  가. 교부액: {fm(project.grant)}원\n"
        f"  나. 집행액: {fm(result.total_spent)}원 (집행률 {rate:.1f}%)\n"
        f"  다. 반납액: {fm(result.remaining)}원 "
        f"(금{korean_amount(result.remaining)}원정)\n"
        f"  라. 반납사유: 계약 절감액 및 잔여 예산\n\n"
        f"※ 자동 생성 초안입니다. 결재 전 담당자 확인이 필요합니다."
    )


# -------------------------------------------------------------------- HTML

_CSS = """
body{font-family:"Pretendard","Apple SD Gothic Neo","Malgun Gothic",sans-serif;
 margin:0;background:#f6f7f9;color:#14171f;font-size:14px;line-height:1.55}
.wrap{max-width:900px;margin:0 auto;padding:24px}
h1{font-size:18px;margin:0 0 4px}.sub{color:#7b8496;font-size:12px;margin-bottom:18px}
.card{background:#fff;border:1px solid #e3e6ec;border-radius:10px;margin-bottom:16px;overflow:hidden}
.card h2{font-size:13px;margin:0;padding:11px 15px;border-bottom:1px solid #eef0f4}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;background:#eef0f4}
.stat{background:#fff;padding:12px 14px}
.stat .lb{font-size:10.5px;color:#7b8496;font-weight:600}
.stat .vl{font-size:17px;font-weight:700;font-variant-numeric:tabular-nums}
.stat .vl.neg{color:#c0362c}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th,td{padding:7px 9px;border-bottom:1px solid #eef0f4;text-align:left}
th{background:#fafbfc;font-size:11px;color:#7b8496}
td.num,th.num{text-align:right;font-variant-numeric:tabular-nums}
.rule{display:flex;gap:9px;padding:9px 12px;border-bottom:1px solid #eef0f4}
.rule .c{font-weight:700;font-size:10.5px;width:32px;flex-shrink:0}
.rule small{display:block;color:#7b8496;font-size:11.5px;margin-top:2px}
.error{background:#fdeeec}.error .c{color:#c0362c}
.warning{background:#fdf3e0}.warning .c{color:#a86500}
.info .c{color:#3d5afe}.ok .c{color:#0f7b4f}
pre{white-space:pre-wrap;font-family:inherit;background:#fafbfc;border:1px solid #e3e6ec;
 border-radius:8px;padding:14px;font-size:12.5px;margin:0}
.note{font-size:11.5px;color:#7b8496;padding:10px 15px}
"""


def html_report(result: Result, project: Project,
                draft: Optional[Draft] = None, quote: Optional[Quote] = None,
                tax: Optional[TaxInvoice] = None) -> str:
    e = html.escape
    s = result.summary()
    rows = "".join(
        f"<tr><td class='num'>{r[0]}</td><td>{e(str(r[1]))}</td><td>{e(str(r[2]))}</td>"
        f"<td>{e(str(r[3]))}</td><td>{e(str(r[4]))}</td>"
        f"<td class='num'>{fm(r[5])}</td><td>{e(str(r[6]))}</td></tr>"
        for r in settlement_rows(project, result, draft, quote))
    findings = "".join(
        f"<div class='rule {f.severity}'><span class='c'>{_MARK[f.severity]} {f.code}</span>"
        f"<span>{e(f.title)}{f'<small>{e(f.detail)}</small>' if f.detail else ''}</span></div>"
        for f in result.sorted_findings)
    refund = refund_draft(project, result)
    period = f"{project.start} ~ {project.end}" if project.start and project.end else ""
    return f"""<!DOCTYPE html><html lang="ko"><head><meta charset="utf-8">
<title>{e(project.name)} 정산 검증 리포트</title><style>{_CSS}</style></head><body><div class="wrap">
<h1>목적사업비 집행내역서 · 검증 리포트</h1>
<div class="sub">{e(project.name)} &nbsp;|&nbsp; 집행기간 {period}</div>

<div class="card"><div class="stats">
  <div class="stat"><div class="lb">교부액</div><div class="vl">{fm(project.grant)}</div></div>
  <div class="stat"><div class="lb">기집행</div><div class="vl">{fm(result.prior_total)}</div></div>
  <div class="stat"><div class="lb">이번 건</div><div class="vl">{fm(result.this_amount)}</div></div>
  <div class="stat"><div class="lb">잔액</div>
    <div class="vl {'neg' if result.remaining < 0 else ''}">{fm(result.remaining)}</div></div>
</div></div>

<div class="card"><h2>검증 결과 — 오류 {s['error']} · 경고 {s['warning']} · 통과 {s['ok']}</h2>
{findings}</div>

<div class="card"><h2>집행내역서</h2>
<table><thead><tr><th class="num">연번</th><th>집행일</th><th>세부항목</th><th>거래처</th>
<th>사업자등록번호</th><th class="num">집행액(원)</th><th>계약방법</th></tr></thead>
<tbody>{rows}</tbody></table>
<div class="note">교부액 {fm(project.grant)} · 집행액 계 {fm(result.total_spent)} ·
 집행잔액 {fm(result.remaining)}</div></div>

{f'<div class="card"><h2>잔액 반납 기안 초안</h2><div style="padding:15px"><pre>{e(refund)}</pre></div></div>' if refund else ''}

<div class="note">본 리포트는 jeongsan 룰 엔진이 생성한 초안입니다.
모든 금액은 원본 서류에서 추출·검산된 값이며, 최종 확정과 결재는 담당자가 수행합니다.</div>
</div></body></html>"""
